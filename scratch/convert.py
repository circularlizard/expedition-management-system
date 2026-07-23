import openpyxl
import json
from datetime import datetime

# Load workbook
wb = openpyxl.load_workbook('docs/Members - 2026-07-23.xlsx')
sheet = wb.active

# Parse rows
members = []
scout_id_counter = 3417257
section_id = 99001

# Map patrol name to id
patrol_map = {
    'S3': '99201',
    'S4': '99202',
    'S5': '99203',
    'S6': '99204',
    'Post S6': '99205'
}

# The reference date for age calculation (based on EMSTestCase '2026-06-13')
ref_date = datetime(2026, 6, 13)

# Row 1 is header grouping, Row 2 is column name, data starts at Row 3
for r in range(3, sheet.max_row + 1):
    fname = sheet.cell(row=r, column=1).value
    lname = sheet.cell(row=r, column=2).value
    dob_val = sheet.cell(row=r, column=3).value
    patrol_val = sheet.cell(row=r, column=4).value
    
    # Member email (index 62, 1-based column 63)
    # Primary Contact 1 email (index 15, 1-based column 16)
    email_val = sheet.cell(row=r, column=63).value
    parent_email_val = sheet.cell(row=r, column=16).value
    
    if not fname or not lname:
        continue
        
    scout_id = str(scout_id_counter)
    scout_id_counter += 1
    
    # Parse DOB
    if isinstance(dob_val, datetime):
        dob_str = dob_val.strftime('%Y-%m-%d')
        # Calculate age
        years = ref_date.year - dob_val.year
        months = ref_date.month - dob_val.month
        if months < 0:
            years -= 1
            months += 12
        age_str = f"{years} / {months}"
    else:
        # Default fallback
        dob_str = "2008-01-01"
        age_str = "18 / 5"
        
    patrol_name = patrol_val if patrol_val else 'S3'
    patrol_id = patrol_map.get(patrol_name, '99201')
    
    member = {
        "scoutid": scout_id,
        "firstname": fname,
        "lastname": lname,
        "patrol": patrol_name,
        "patrolid": patrol_id,
        "sectionid": section_id,
        "dob": dob_str,
        "age": age_str,
        "_filterString": f"{fname.lower()} {lname.lower()}",
        # extra details for detailed map
        "email": email_val if email_val else f"scout.{scout_id}@example-ems.test",
        "parent_email": parent_email_val if parent_email_val else f"parent.{scout_id}@example-ems.test"
    }
    members.append(member)

print(f"Parsed {len(members)} members.")
if len(members) > 0:
    print("Example:", json.dumps(members[0], indent=4))
